import os
import xml.etree.ElementTree as ET
import openpyxl
import pandas as pd

# Set the directory you want to start from
rootDir = r"C:\Users\MR\Documents\KSB\6875\XML"

# Create an empty list to store the file names
file_names = []


# Walk through all the directories and subdirectories
for dirName, subdirList, fileList in os.walk(rootDir):
    # Print the current directory
    print("Found directory: %s" % dirName)

    # Iterate over the list of files
    for fname in fileList:
        # Append the file name to the list

        file_names.append(r"%s" % fname)

# Compare the file names
for i in range(len(file_names)):
    for j in range(i + 1, len(file_names)):

        f1name = file_names[i]
        f1name = f1name[:25]
        f2name = file_names[j]
        f2name = f2name[:25]
        # if file_names[i][:26] == file_names[j]:
        if f1name == f2name:
            print("Process Started for %s" % file_names[i])
            # maintain dir here also
            f1 = r"C:\Users\MR\Documents\KSB\6875\XML\%s" % file_names[i]
            f2 = r"C:\Users\MR\Documents\KSB\6875\XML\%s" % file_names[j]
            tree1 = ET.parse(f1)
            tree2 = ET.parse(f2)
            root1 = tree1.getroot()
            root2 = tree2.getroot()

            # create a new workbook and worksheet
            workbook = openpyxl.Workbook()
            worksheet = workbook.create_sheet(title="Old")

            def write_data(worksheet, row, data):
                for i, item in enumerate(data):
                    worksheet.cell(row=row, column=i + 1).value = item

            # recursive function to process elements and child nodes
            def process_element(element, row, level=0):
                data = [level, element.tag, element.text]
                for name, value in element.attrib.items():
                    data.append(f"{name}: {value}")
                write_data(worksheet, row, data)
                row += 1
                for child in element:
                    row = process_element(child, row, level + 1)
                    row += 1
                    for grandchild in child:
                        row = process_element(grandchild, row, level + 1)
                return row

            # start processing at the root element
            process_element(root1, 1)
            worksheet = workbook.create_sheet(title="New")
            process_element(root2, 1)
            target = r"C:\Users\MR\Documents\KSB\6875\Excel\%s.xlsx" % f1name
            # save the workbook
            workbook.save(target)
